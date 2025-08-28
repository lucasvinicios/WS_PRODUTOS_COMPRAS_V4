from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

def realizar_analise_reversa_acumulativa(file_path):
    """
    Realiza a análise de preços de produtos, encontrando o item mais barato
    em cada mercado para a cesta de compras.
    """
    mercados_data = {}

    excel_file = pd.ExcelFile(file_path)
    sheet_names = excel_file.sheet_names
    
    for sheet_name in sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # A lógica para extrair os dados de cada aba permanece a mesma
        df.set_index('Produto', inplace=True)
        nome_mercado_curto = df.columns[0]
        print("Aqui")
        print(df)
        frete = df.loc['Frete', nome_mercado_curto]
        
        df_produtos = df.drop(index=['Frete', 'Valor Mínimo', 'Valor Total', 'Total Baratos + Frete'], errors='ignore')

        mercados_data[sheet_name] = {
            'df': df_produtos,
            'frete': frete,
            'nome_curto': nome_mercado_curto
        }

    resultados_finais = []
    
    custos_totais = pd.DataFrame()
    for nome_arquivo_completo, data in mercados_data.items():
        df_temp = data['df'].copy()
        df_temp['Custo Total'] = df_temp.iloc[:, 0] + data['frete']
        custos_totais[nome_arquivo_completo] = df_temp['Custo Total']
    
    for produto in custos_totais.index:
        custos_ordenados = custos_totais.loc[produto].dropna().sort_values()
        
        if not custos_ordenados.empty:
            mercado_mais_barato_key = custos_ordenados.index[0]
            preco_produto = mercados_data[mercado_mais_barato_key]['df'].loc[produto].iloc[0]
            frete = mercados_data[mercado_mais_barato_key]['frete']
            custo_total = preco_produto + frete
            
            resultados_finais.append({
                'Produto': produto,
                'Supermercado Recomendado': mercado_mais_barato_key,
                'Preço do Produto': preco_produto,
                'Frete': frete,
                'Custo Total': custo_total,
            })
            
    return pd.DataFrame(resultados_finais).sort_values(by='Produto'), mercados_data

def gerar_aba_comparacao_dinamica(writer, mercados_data, analise_final_df):
    """
    Gera uma aba de comparação para os três supermercados mais baratos
    com a fórmula de economia explícita do usuário.
    """
    # 1. Identifica os 3 supermercados com a menor cesta total
    custos_totais_mercados = {}
    for nome_completo, data in mercados_data.items():
        custo_total = data['df'].iloc[:, 0].sum() + data['frete']
        custos_totais_mercados[nome_completo] = custo_total
    
    top_3_mercados = sorted(custos_totais_mercados.items(), key=lambda item: item[1])[:3]
    top_3_nomes_completos = [item[0] for item in top_3_mercados]
    
    print(top_3_nomes_completos)

    nomes_sanitizados = []

    for nome_sanitizado in top_3_nomes_completos:
        nomes_sanitizados.append(nome_sanitizado.split(" ")[-1])

    print(nomes_sanitizados)

    # 2. Filtra os dados apenas para esses 3 supermercados
    mercados_selecionados = {k: v for k, v in mercados_data.items() if k in top_3_nomes_completos}
    
    # 3. Coleta os preços e fretes
    precos_comparativos = pd.DataFrame()
    fretes = {}
    for nome_completo, data in mercados_selecionados.items():
        precos_comparativos[data['nome_curto']] = data['df'].iloc[:, 0]
        fretes[data['nome_curto']] = data['frete']
    
    precos_comparativos.loc['Frete'] = pd.Series(fretes)

    # 4. Calcula o "Total Economizado" usando a fórmula explícita
    economias_dinamica = {}

    # DataFrame com todos os preços de todos os mercados (somente produtos)
    todos_os_precos = pd.DataFrame()
    for nome_completo, data in mercados_data.items():
        todos_os_precos[data['nome_curto']] = data['df'].iloc[:, 0]

    print(todos_os_precos)

    todos_os_precos = todos_os_precos[nomes_sanitizados]

    # Para cada um dos 3 supermercados que estamos comparando
    for mercado_curto_atual in precos_comparativos.columns:
        frete_mercado_atual = fretes[mercado_curto_atual]
        soma_das_diferencas = 0

        # Itera sobre cada produto
        for produto in todos_os_precos.index:
            # Pega todos os preços para o produto atual, remove NaNs e ordena
            precos_produto_ordenados = todos_os_precos.loc[produto].dropna().sort_values()[:3]
            
            
            # Precisa de pelo menos 2 preços para calcular a diferença
            if len(precos_produto_ordenados) >= 2:
                mercado_mais_barato_curto = precos_produto_ordenados.index[0]
                # print(mercado_mais_barato_curto)
                # print(mercado_curto_atual)
                # Verifica se o supermercado atual é o mais barato para este produto
                if mercado_mais_barato_curto == mercado_curto_atual:

                    preco_mais_barato = precos_produto_ordenados.iloc[0]
                    # print(preco_mais_barato)
                    preco_segundo_mais_barato = precos_produto_ordenados.iloc[1]
                    # print(preco_segundo_mais_barato)
                    soma_das_diferencas += (preco_segundo_mais_barato - preco_mais_barato)

                    if 'TENDA' in mercado_curto_atual:
                        print(produto, preco_mais_barato, preco_segundo_mais_barato, soma_das_diferencas)
                    # print(soma_das_diferencas)

        # Aplica a fórmula final conforme solicitado: Frete + Soma das Diferenças
        economias_dinamica[mercado_curto_atual] = soma_das_diferencas - frete_mercado_atual

    # print(economias_dinamica)
    
    precos_comparativos.loc['Total Economizado'] = pd.Series(economias_dinamica)

    # 5. Escreve os dados na planilha
    nome_aba = 'Comparativo Dinâmico'
    precos_comparativos.index.name = 'Produto'
    precos_comparativos.to_excel(writer, sheet_name=nome_aba, index=True)
    
    # 6. Aplica a formatação condicional e de valores
    workbook = writer.book
    worksheet = workbook[nome_aba]

    verde_fill = PatternFill(start_color='63BE7B', end_color='63BE7B', fill_type='solid') # verde
    azul_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid') # azul
    
    num_mercados = len(precos_comparativos.columns)
    
    for row_idx, produto in enumerate(precos_comparativos.index, start=2):
        if produto in ['Frete', 'Total Economizado']:
            continue
            
        valores = [worksheet.cell(row=row_idx, column=i).value for i in range(2, num_mercados + 2)]
        
        valores_numericos = [v for v in valores if isinstance(v, (int, float))]
        
        if valores_numericos:
            valores_ordenados = sorted(valores_numericos)
            
            min_val = valores_ordenados[0]
            for col_idx in range(2, num_mercados + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value == min_val:
                    cell.fill = verde_fill
            
            if len(valores_ordenados) >= 2:
                second_min_val = valores_ordenados[1]
                for col_idx in range(2, num_mercados + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value == second_min_val and cell.fill != verde_fill:
                        cell.fill = azul_fill
                        
    for col in range(2, num_mercados + 2):
        for row in range(2, len(precos_comparativos.index) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.number_format = '"R$" #,##0.00'
            
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    print("Aba 'Comparativo Dinâmico' gerada com sucesso e formatada.")

def gerar_excel_final(analise_final_df, mercados_data, nome_arquivo='analise_supermercados_final.xlsx'):
    """
    Gera um arquivo Excel com abas para cada supermercado recomendado e um resumo.
    """
    custos_totais_com_nomes_curtos = pd.DataFrame()
    for nome_arquivo_completo, data in mercados_data.items():
        nome_mercado_curto = data['nome_curto']
        df_temp = data['df'].copy()
        df_temp['Custo Total'] = df_temp.iloc[:, 0] - data['frete']
        custos_totais_com_nomes_curtos[nome_mercado_curto] = df_temp['Custo Total']
    
    economias = {}
    for produto in custos_totais_com_nomes_curtos.index:
        custos_ordenados = custos_totais_com_nomes_curtos.loc[produto].sort_values()
        custo_melhor = custos_ordenados.iloc[0]
        try:
            custo_segundo_melhor = custos_ordenados.iloc[1]
            economia_calculada = custo_segundo_melhor - custo_melhor
            economias[produto] = economia_calculada
        except IndexError:
            economias[produto] = 0

    mercados_recomendados_keys = analise_final_df['Supermercado Recomendado'].unique()
    
    mercados_recomendados_ordenados = sorted(
        mercados_recomendados_keys,
        key=lambda x: int(x.split(' ')[1]) if 'TOP' in x else float('inf')
    )

    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        
        gerar_aba_comparacao_dinamica(writer, mercados_data, analise_final_df)

        # for nome_mercado_completo in mercados_recomendados_ordenados:
        #     data = mercados_data[nome_mercado_completo]
        #     df_mercado = data['df'].copy()
        #     frete = data['frete']
            
        #     df_mercado['Custo Total (c/ Frete)'] = df_mercado.iloc[:, 0] + frete
        #     df_mercado['Economia (vs 2º Melhor)'] = 0.00
            
        #     for _, row in analise_final_df[analise_final_df['Supermercado Recomendado'] == nome_mercado_completo].iterrows():
        #         produto = row['Produto']
        #         df_mercado.loc[produto, 'Economia (vs 2º Melhor)'] = economias.get(produto, 0)
            
        #     total_economizado = df_mercado['Economia (vs 2º Melhor)'].sum()

        #     df_frete = pd.DataFrame({'Frete': frete, 'Custo Total (c/ Frete)': ''}, index=['Frete'])
        #     df_frete.index.name = 'Produto'
            
        #     df_total = pd.DataFrame({'Economia (vs 2º Melhor)': total_economizado}, index=['Total Economizado'])
        #     df_total.index.name = 'Produto'

        #     df_final = df_mercado[['Custo Total (c/ Frete)', 'Economia (vs 2º Melhor)']].copy()
        #     df_final.insert(0, 'Preço Unitário', df_mercado.iloc[:, 0])
            
        #     df_final = pd.concat([df_final, df_frete, df_total])
            
        #     df_final = df_final.drop(columns=['Frete'], errors='ignore')
            
        #     nome_aba = nome_mercado_completo.replace(' ', ' - ', 1)
            
        #     df_final.to_excel(writer, sheet_name=nome_aba, index=True)
        #     print(f"Aba '{nome_aba}' gerada com sucesso. Total economizado: R$ {total_economizado:.2f}")

# Dados de exemplo fornecidos
# arquivos_do_excel = {
#     "TOP 1 TENDAATACADO": """Produto,TENDAATACADO
# Arroz 5kg,17.4
# Feijão 1kg,3.55
# Macarrão 500g,2.59
# Óleo 900ml,6.19
# Açúcar 5kg,18.29
# Leite Integral 1L,5.45
# Pão de Forma 500g,5.59
# Café 500g,21.99
# Detergente 500ml,1.8
# Sabão em Pó 1kg,9.25
# Papel Higiênico,8.2
# Creme Dental 70g,2.6
# Água Sanitária,3.45
# Sabonete,1.05
# Fio Dental,5.15
# Molho de Tomate,1.25
# Azeite 500ml,32.89
# Farinha de Trigo 1kg,2.99
# Queijo 200g,9.49
# Creme de Leite 200g,2.25
# Frete,14.9
# """,
#     "TOP 2 BOASUPERMERCADO": """Produto,BOASUPERMERCADO
# Arroz 5kg,20.9
# Feijão 1kg,4.69
# Macarrão 500g,3.49
# Óleo 900ml,7.89
# Açúcar 5kg,22.79
# Leite Integral 1L,5.19
# Pão de Forma 500g,6.59
# Café 500g,28.79
# Detergente 500ml,1.85
# Sabão em Pó 1kg,4.19
# Papel Higiênico,4.09
# Creme Dental 70g,1.55
# Água Sanitária,4.99
# Sabonete,1.55
# Fio Dental,5.99
# Molho de Tomate,1.79
# Azeite 500ml,39.9
# Farinha de Trigo 1kg,3.49
# Queijo 200g,10.99
# Creme de Leite 200g,3.45
# Frete,15
# """,
#     "TOP 3 TAUSTE": """Produto,TAUSTE
# Arroz 5kg,19.89
# Feijão 1kg,5.77
# Macarrão 500g,3.98
# Óleo 900ml,6.97
# Açúcar 5kg,16.89
# Leite Integral 1L,4.59
# Pão de Forma 500g,6.86
# Café 500g,27.69
# Detergente 500ml,1.79
# Sabão em Pó 1kg,4.59
# Papel Higiênico,6.37
# Creme Dental 70g,1.79
# Água Sanitária,4.59
# Sabonete,1.19
# Fio Dental,7.69
# Molho de Tomate,1.78
# Azeite 500ml,38.89
# Farinha de Trigo 1kg,3.29
# Queijo 200g,8.77
# Creme de Leite 200g,2.98
# Frete,14.9
# """,
#     "TOP 4 CONFIANCA": """Produto,CONFIANCA
# Arroz 5kg,19.89
# Feijão 1kg,5.77
# Macarrão 500g,2.98
# Óleo 900ml,6.98
# Açúcar 5kg,16.79
# Leite Integral 1L,5.19
# Pão de Forma 500g,6.7
# Café 500g,27.59
# Detergente 500ml,1.98
# Sabão em Pó 1kg,9.78
# Papel Higiênico,4.79
# Creme Dental 70g,1.79
# Água Sanitária,4.79
# Sabonete,1.77
# Fio Dental,9.68
# Molho de Tomate,1.15
# Azeite 500ml,38.9
# Farinha de Trigo 1kg,3.39
# Queijo 200g,8.98
# Creme de Leite 200g,2.79
# Frete,18.9
# """,
#     "TOP 5 BARBOSA": """Produto,BARBOSA
# Arroz 5kg,19.99
# Feijão 1kg,5.99
# Macarrão 500g,2.79
# Óleo 900ml,7.49
# Açúcar 5kg,24.95
# Leite Integral 1L,5.49
# Pão de Forma 500g,6.49
# Café 500g,27.99
# Detergente 500ml,2.19
# Sabão em Pó 1kg,7.99
# Papel Higiênico,7.99
# Creme Dental 70g,2.99
# Água Sanitária,4.49
# Sabonete,0
# Fio Dental,8.99
# Molho de Tomate,1.49
# Azeite 500ml,37.99
# Farinha de Trigo 1kg,3.99
# Queijo 200g,8.99
# Creme de Leite 200g,3.49
# Frete,20.9
# """,
#     "TOP 6 COOPSUPERMERCADO": """Produto,COOPSUPERMERCADO
# Arroz 5kg,25.99
# Feijão 1kg,5.99
# Macarrão 500g,3.59
# Óleo 900ml,7.49
# Açúcar 5kg,24.99
# Leite Integral 1L,5.99
# Pão de Forma 500g,6.79
# Café 500g,23.79
# Detergente 500ml,1.79
# Sabão em Pó 1kg,19.29
# Papel Higiênico,9.29
# Creme Dental 70g,3.49
# Água Sanitária,4.99
# Sabonete,1.59
# Fio Dental,15.19
# Molho de Tomate,1.69
# Azeite 500ml,42.49
# Farinha de Trigo 1kg,4.29
# Queijo 200g,11.99
# Creme de Leite 200g,3.39
# Frete,15
# """
# }

arquivos_do_excel = f"Resultados\\analise_supermercados_2025-08-27.xlsx"

try:
    # 1. Executa a análise lendo diretamente do arquivo Excel
    analise_completa, mercados_dados = realizar_analise_reversa_acumulativa(arquivos_do_excel)

    # 2. Gera o novo arquivo Excel
    gerar_excel_final(analise_completa, mercados_dados)
    
except FileNotFoundError:
    print(f"Erro: O arquivo '{arquivos_do_excel}' não foi encontrado.")
except Exception as e:
    print(f"Ocorreu um erro: {e}. Linha de erro: {e.__traceback__.tb_lineno}")